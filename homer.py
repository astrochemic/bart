import pandas as pd

from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from db_queries import _get_platform_active, _get_platform_schedule, _get_redshift_transactions
from helper import make_local_filename


class Country(object):
    def __init__(self, country, start_date, end_date, broadcast_config, verbose=True):
        self.country = country
        self.start_date = start_date
        self.end_date = end_date
        self.broadcast_config = broadcast_config[
            broadcast_config['country_code'] == country
        ].drop(columns='country_code')
        self.verbose = verbose
        self.df_active = pd.DataFrame()
        self.df_active_per_service = pd.DataFrame()
        self.df_plat_trx = pd.DataFrame()
        self.df_plat_trx_per_service = pd.DataFrame()
        self.df_plat_trx_per_user = pd.DataFrame()
        self.df_red_trx = pd.DataFrame()
        self.df_red_trx_per_service = pd.DataFrame()
        self.df_active_with_trx = pd.DataFrame()
        self.results = pd.DataFrame()

    def get_active_users(self):
        if self.verbose: print('Finding active users in {}...'.format(self.country))
        self.df_active = _get_platform_active(self.country, self.start_date, self.end_date)
        if self.df_active.empty:
            return
        self.df_schedules = _get_platform_schedule(self.country, self.start_date, self.end_date)
        self.select_unique_schedules()
        self.df_active = pd.merge(
            self.df_active,
            self.df_schedules[['serviceid', 'operator_code', 'tariff', 'billing_days']],
            how='left',
            on=['serviceid', 'operator_code']
        )
        if self.df_active['billing_days'].notnull().sum() > 0:
            self.df_active['frequency'] = self.df_active['billing_days'].str.split(',').str.len().fillna(-1).astype(int)
        else:
            self.df_active['frequency'] = -1
        self.df_active['tariff'].fillna(-1, inplace=True)
        self.df_active['billing_days'].fillna('', inplace=True)

        if self.country == 'MY':
            self.df_active['service_identifier2'] = (
                self.df_active['service_identifier2'].str.replace('ON ', '')
            )

        columns = ['platform', 'gateway', 'operator_code',
                   'service_identifier1', 'service_identifier2', 'frequency']
        df = self.df_active[columns].copy()
        df['active_users'] = 1
        self.df_active_per_service = df.groupby(
            columns, as_index=False
        ).agg('sum').sort_values(columns).reset_index(drop=True)

        self.add_broadcast_config()

    def select_unique_schedules(self):
        if self.df_schedules.empty or 'serviceid' not in self.df_schedules:
            return
        df_unique_schedules = pd.DataFrame()
        serviceids = self.df_schedules['serviceid'].unique()
        for serviceid in serviceids:
            df_temp = pd.DataFrame()
            df1 = self.df_schedules[self.df_schedules['serviceid'] == serviceid]
            for i, row in df1.iterrows():
                operator_codes = row['operator_code'].split(',')
                for operator_code in operator_codes:
                    row['operator_code'] = operator_code
                    df_temp = df_temp.append(row, sort=True)
            operator_codes = df_temp['operator_code'].unique()
            for operator_code in operator_codes:
                df2 = df_temp[df_temp['operator_code'] == operator_code]
                if len(df2) == 1:
                    schedule = df2.iloc[0]
                else:
                    idx = df2['schedule_status'] == 'A'
                    if idx.sum() == 1:
                        schedule = df2[idx].iloc[0]
                    elif idx.sum() == 0:
                        schedule = df2.sort_values('updatedate').iloc[-1]
                    else:
                        schedule = df2[idx].sort_values('updatedate').iloc[-1]
                df_unique_schedules = df_unique_schedules.append(schedule, sort=True)
        self.df_schedules = df_unique_schedules.copy().reset_index(drop=True)

    def add_broadcast_config(self):
        if 'match_score' in self.broadcast_config.columns:
            self.broadcast_config = self.broadcast_config.drop(columns='match_score')
        drilldown_columns = ['gateway', 'operator_code',
                             'service_identifier1', 'service_identifier2']
        broadcast_info = {col: [] for col in self.broadcast_config.columns
                          if col not in drilldown_columns}
        for i, row in self.df_active_per_service.iterrows():
            match_scores = []
            for j, config in self.broadcast_config.iterrows():
                mismatch = False
                score = 0
                for k, col in enumerate(drilldown_columns):
                    if config[col] != '*' and config[col] != row[col]:
                        mismatch = True
                    elif config[col] == row[col]:
                        score += 2 ** k
                if mismatch:
                    match_scores.append(0)
                    continue
                match_scores.append(score)
            self.broadcast_config['match_score'] = match_scores
            if self.broadcast_config['match_score'].max() > 0:
                iconf = self.broadcast_config['match_score'].idxmax()
                for col in broadcast_info:
                    broadcast_info[col].append(self.broadcast_config.loc[iconf, col])
            else:
                for col in broadcast_info:
                    broadcast_info[col].append('?')
        for col in broadcast_info:
            self.df_active_per_service[col] = broadcast_info[col]
        self.df_active_per_service['muted'] = (pd.to_numeric(
            self.df_active_per_service['minimum_expected_transactions'], errors='coerce'
        ) < 0) & (pd.to_numeric(
            self.df_active_per_service['maximum_expected_transactions'], errors='coerce'
        ) < 0)

    def get_redshift_transactions(self):
        if self.verbose: print('Finding transactions for {}...'.format(self.country))
        self.df_red_trx_per_user = _get_redshift_transactions(
            self.country, self.start_date, self.end_date
        )

    def print_summary(self):
        if self.df_active.empty:
            return

        print('\nActive users ({}):'.format(self.country))
        print(self.df_active[['platform', 'msisdn']].groupby(
            'platform', as_index=False
        ).agg('count'))
        print()

        print('Total transactions in Redshift ({}):'.format(self.country))
        print(self.df_red_trx_per_user[['platform', 'total_transactions']].groupby(
            'platform', as_index=False
        ).agg('sum'))
        print()

        print('Total transactions matched to active users ({}):'.format(self.country))
        print(self.df_active_with_trx[['platform', 'total_transactions']].groupby(
            'platform', as_index=False
        ).agg('sum'))
        print()

    def run_analysis(self):
        if self.verbose: print('Analyzing broadcasts in {}...'.format(self.country))
        if self.df_active.empty:
            return

        self.results = pd.DataFrame()
        for i, service in self.df_active_per_service.iterrows():
            if service['muted']:
                continue

            df_this_service = get_active_for_service(service, self.df_active_with_trx)
            if len(df_this_service) == 0:
                print(i)

            df2 = df_this_service[['total_transactions', 'users']].groupby(
                'total_transactions', as_index=False
            ).agg('sum')
            these_results = pd.DataFrame(service).transpose()
            if 0 not in df2['total_transactions'].values:
                these_results[0] = 0
            for j, row in df2.iterrows():
                these_results[row['total_transactions']] = row['users']
            self.results = self.results.append(these_results, sort=True)

        if self.results.empty:
            return

        self.df_active_per_service.drop(columns='muted', inplace=True)
        self.results.drop(columns='muted', inplace=True)

        if len(self.results) > 0:
            trx_counts = sorted(
                set(self.results.columns) - set(self.df_active_per_service.keys())
            )
            for c in trx_counts:
                self.results[c] = self.results[c].fillna(0).astype(int)
            columns = list(self.df_active_per_service.keys()) + trx_counts
            self.results = self.results[columns]

    def set_active_with_trx(self):
        if self.verbose:
            print('Linking transactions to active users in {}...'.format(self.country))
        if self.df_active.empty:
            return

        trx_columns = ['total_transactions', 'delivered_transactions']

        # Match transactions to users with a rockman_id
        idx = self.df_active['rockman_id'].notnull()
        if idx.sum() > 0:
            self.df_active_with_trx = pd.merge(
                self.df_active[idx],
                self.df_red_trx_per_user[['rockman_id'] + trx_columns],
                how='left',
                on='rockman_id'
            )
        else:
            self.df_active_with_trx = pd.DataFrame(
                columns=list(self.df_active.columns)+trx_columns
            )

        # Match transactions to users without a rockman_id
        idx = self.df_active['rockman_id'].isnull()
        columns = ['msisdn', 'platform', 'service_identifier2']
        if self.country == 'BE':
            columns = ['msisdn']
        elif self.country == 'IQ':
            columns = ['msisdn', 'service_identifier2']
        self.df_active_with_trx = self.df_active_with_trx.append(pd.merge(
            self.df_active[idx],
            self.df_red_trx_per_user[columns + trx_columns],
            how='left',
            on=columns
        ), sort=True)

        for col in trx_columns:
            self.df_active_with_trx[col] = (
                self.df_active_with_trx[col].fillna(0).astype(int)
            )
        self.df_active_with_trx['users'] = 1

    def write_csv(self):
        if self.verbose: print('Writing results to CSV for {}...'.format(self.country))
        if self.df_active.empty or self.results.empty:
            return
        output_file = make_local_filename(self.start_date, self.end_date, self.country, 'csv')
        self.results.to_csv(output_file, index=False)

    def write_excel(self, wb):
        if self.verbose: print('Writing results to XLSX for {}...'.format(self.country))
        if self.df_active.empty or self.results.empty:
            return

        df = self.results.drop(columns='frequency')

        # Dump results into basic sheet
        ws = wb.create_sheet(self.country)
        ws.append([])
        trxcol1 = len(self.df_active_per_service.columns) - 1
        ntrxcol = len(df.columns) - trxcol1
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Insert column with numbers of users with at least one transaction
        ws.cell(row=2, column=trxcol1+1).value = 'Users without transactions'
        ws.insert_cols(trxcol1+2)
        ws.cell(row=2, column=trxcol1+2).value = 'Users with transactions'
        for row in ws.iter_rows(min_row=3, min_col=trxcol1+2, max_col=trxcol1+2):
            for cell in row:
                cell.value = '=SUM({}{}:{}{})'.format(
                    get_column_letter(trxcol1+3), cell.row,
                    get_column_letter(trxcol1+1+ntrxcol), cell.row,
                ) if ntrxcol > 1 else 0

        # Move 'active_users' column to the right by two columns
        icol = list(df.columns).index('active_users') + 1
        ws.insert_cols(trxcol1+1)
        for row in ws.iter_rows(min_row=2, min_col=trxcol1+1, max_col=trxcol1+1):
            for cell in row:
                cell.value = ws.cell(row=cell.row, column=icol).value
        ws.delete_cols(icol, 1)

        # Add/customize column headers
        if ntrxcol > 1:
            ws.cell(row=1, column=trxcol1+3).value = (
                '# of users with a given number of transactions this week'
            )
            ws.cell(row=1, column=trxcol1+3).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=1, start_column=trxcol1+3,
                           end_row=1, end_column=trxcol1+ntrxcol+1)
        icol = list(df.columns).index('handled_by') - 1
        for cell in ws[2][icol:trxcol1+2]:
            cell.value = cell.value.capitalize().replace('_', ' ').replace(
                ' dns', ' DNs')

        # Set column widths and row heights
        widths = [56, 100, 123, 109, 109, 54, 41, 62, 46, 73, 73, 52, 73, 73]
        for j, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(j+1)].width = w / 6.
        ws.row_dimensions[2].height = 48

        # Apply cell formatting
        for cell in ws[2][0:]:
            cell.alignment = Alignment(wrap_text=True)
        for row in ws.iter_rows(min_row=3, max_col=trxcol1-1):
            for cell in row:
                cell.number_format = numbers.FORMAT_TEXT

        # Set font size
        for row in ws:
            for cell in row:
                cell.font = Font(size=12)

def get_active_for_service(service, df):
    keys = ['platform', 'gateway', 'operator_code',
            'service_identifier1', 'service_identifier2', 'frequency']
    idx = df['accountid'].notnull()
    for key in keys:
        if key in service.keys():
            idx = idx & (df[key] == service[key])
    return df[idx].copy().reset_index(drop=True)


def get_active_from_excel(excel_string, df):
    keys = ['platform', 'gateway', 'operator_code',
            'service_identifier1', 'service_identifier2', 'frequency']
    values = excel_string.split('\t')
    service = pd.Series(dict(zip(keys[:len(values)], values)))
    if 'frequency' in service:
        service['frequency'] = int(service['frequency'])
    return get_active_for_service(service, df)
